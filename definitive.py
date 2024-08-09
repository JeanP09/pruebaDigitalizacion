import os
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from PyPDF2 import PdfReader
from datetime import datetime
import win32com.client as win32

directorio = r'C:\Users\user\OneDrive\PruebaDigitalizacion\Prueba1'
ruta_excel_existente = r'C:\Users\user\Documents\pruebaDigitalizacion\FormatoIndiceElectronico.xlsm'
ruta_excel_nueva = r'C:\Users\user\Desktop\00IndiceElectronicoC001Principall.xlsm'

archivos_info = []

def limpiar_nombre_archivo(nombre_archivo):
    nombre_sin_extension = os.path.splitext(nombre_archivo)[0]
    nombre_limpio = re.sub(r'^\d{3}', '', nombre_sin_extension)
    nombre_limpio = nombre_limpio.replace(' ', '')
    return nombre_limpio

def obtener_tipo_formato(archivo):
    extension = os.path.splitext(archivo)[1].lower()
    formatos = {
        '.pdf': '.pdf',
        '.jpeg': 'JPEG',
        '.jpg': 'JPG',
        '.jpe': 'JPEG',
        '.jpg2': 'JPEG2000',
        '.mp3': 'MP3',
        '.wav': 'WAVE',
        '.mpg': 'MPEG-1',
        '.mp1': 'MPEG-1',
        '.mp2': 'MPEG-2',
        '.mp3': 'MPEG-1',
        '.m1v': 'MPEG-1',
        '.m1a': 'MPEG-1',
        '.m2a': 'MPEG-2',
        '.mpa': 'MPEG',
        '.mpv': 'MPEG-4',
        '.mpeg': 'MPEG-1',
        '.m4v': 'MPEG-4'
    }
    return formatos.get(extension, 'Desconocido')

contador = 1

try:
    excel_app = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel_app.Workbooks.Open(ruta_excel_existente)
    excel_app.Visible = False

    num_archivos = len([f for f in os.listdir(directorio) if os.path.isfile(os.path.join(directorio, f))])
    for _ in range(num_archivos):
        excel_app.Run('Macro1InsertarFila')

    workbook.Save()
    workbook.Close()
except Exception as e:
    print(f"Error al manipular Excel: {e}")
finally:
    try:
        excel_app.Quit()
    except Exception as quit_exception:
        print(f"Error al cerrar Excel: {quit_exception}")

try:
    archivos_info = []
    for archivo in os.listdir(directorio):
        ruta_archivo = os.path.join(directorio, archivo)
        if os.path.isfile(ruta_archivo):
            nombre_archivo_original = archivo
            nombre_archivo_limpio = limpiar_nombre_archivo(nombre_archivo_original)
            fecha_creacion = datetime.fromtimestamp(os.path.getctime(ruta_archivo)).strftime('%d/%m/%Y')
            fecha_modificacion = datetime.fromtimestamp(os.path.getmtime(ruta_archivo)).strftime('%d/%m/%Y')
            size_file = round(os.path.getsize(ruta_archivo) / 1024)
            size_file = f"{size_file} KB"
            
            if archivo.lower().endswith('.pdf'):
                try:
                    reader = PdfReader(ruta_archivo)
                    num_paginas = len(reader.pages)
                except Exception as e:
                    print(f"Error al leer el PDF {archivo}: {e}")
                    num_paginas = 'Error al leer'
            else:
                num_paginas = 1

            archivos_info.append({
                'Número secuencial': contador,
                'Nombre del archivo': nombre_archivo_limpio,
                'Fecha de creación': fecha_creacion,
                'Fecha de modificación': fecha_modificacion,
                'Número de páginas': num_paginas,
                'Tamaño del archivo': size_file,
                'Tipo de formato': obtener_tipo_formato(archivo),
                'Origen del archivo': 'Digitalizado'
            })
            contador += 1

    wb = load_workbook(ruta_excel_existente, keep_vba=True)
    ws = wb.active

    def write_to_cell(sheet, cell, value):
        if isinstance(sheet[cell], MergedCell):
            for rango in sheet.merged_cells.ranges:
                if cell in rango:
                    sheet.unmerge_cells(str(rango))
                    break
        sheet[cell] = value

    fila_inicial = 11
    for i, archivo_info in enumerate(archivos_info):
        fila_actual = fila_inicial + i
        write_to_cell(ws, f'A{fila_actual}', archivo_info['Nombre del archivo'])
        write_to_cell(ws, f'B{fila_actual}', archivo_info['Fecha de creación'])
        write_to_cell(ws, f'C{fila_actual}', archivo_info['Fecha de modificación'])
        write_to_cell(ws, f'D{fila_actual}', archivo_info['Número secuencial'])
        write_to_cell(ws, f'E{fila_actual}', archivo_info['Número de páginas'])
        write_to_cell(ws, f'I{fila_actual}', archivo_info['Tamaño del archivo'])
        write_to_cell(ws, f'H{fila_actual}', archivo_info['Tipo de formato'])
        write_to_cell(ws, f'J{fila_actual}', archivo_info['Origen del archivo'])

    wb.save(ruta_excel_nueva)

    print(f'Información de archivos guardada en {ruta_excel_nueva}')
except Exception as e:
    print(f"Error al procesar archivos: {e}")